#!/usr/bin/env python3
"""Generate the publisher figure/alt-text spreadsheet from the built EPUB.

Usage: gen-figure-spreadsheet.py <unzipped-epub-dir> <template.xlsx> <out.xlsx>

One row per numbered figure caption (figcaption or tex4ht div.caption with a
"Figure N.M" id), in spine/document order. Images are associated with the
next numbered caption that follows them (all imgs since the previous numbered
caption). License / reference info comes from the two 00_METADATA.bib files
plus the special-case rules documented inline.
"""
import re, sys, os, glob, html
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment

EPUB_DIR, TEMPLATE, OUT = sys.argv[1], sys.argv[2], sys.argv[3]
REPO = Path(__file__).resolve().parent.parent.parent
BASE = REPO / "Intro-Math-Programming" / "baseText"

# rotated table headers (blank/decorative picture renders) and the cover
SKIP_IMGS = {"cover.png"} | {f"book1-epub{i}x.png" for i in range(5)}
# genuine linear-algebra appendix picture-environment sketches (base names)
LYRYX_IMGS = {f"book1-epub{i}x" for i in range(5, 12)}

# ---------------------------------------------------------------- bibs
def parse_bib(path):
    txt = "\n".join(l for l in open(path).read().splitlines()
                    if not l.lstrip().startswith("%"))
    out = {}
    for m in re.finditer(r"@\w+\{([^,]+),", txt):
        key = m.group(1).strip(); start = m.end(); d = 1; i = start
        while i < len(txt) and d > 0:
            if txt[i] == "{": d += 1
            elif txt[i] == "}": d -= 1
            i += 1
        body = txt[start:i - 1]; fields = {}
        for fm in re.finditer(r"(\w+)\s*=\s*\{", body):
            j = fm.end(); dd = 1; k = j
            while k < len(body) and dd > 0:
                if body[k] == "{": dd += 1
                elif body[k] == "}": dd -= 1
                k += 1
            fields[fm.group(1).lower()] = " ".join(body[j:k - 1].split())
        out[key] = fields
    return out

BIBS = {}
for bp in ["optimization/figures/figures-static/00_METADATA.bib",
           "optimization/figures/figures-source/00_METADATA.bib"]:
    BIBS.update(parse_bib(BASE / bp))

def bib_entry(base):
    for cand in (base, base + ".png", base + ".jpg", base + ".pdf",
                 base + ".JPG", "tikz/" + base, "tikz/" + base + ".pdf"):
        if cand in BIBS:
            return BIBS[cand]
    return None

OWN = re.compile(r"Hildebrand|Fravel")
BOOK_LIC = "CC BY-SA 4.0 (original figure created for this book)"

def clean_author(a):
    a = re.sub(r"\\href\{[^}]*\}\{([^}]*)\}", r"\1", a)
    a = re.sub(r"\\\w+", "", a).replace("{", "").replace("}", "")
    return " ".join(a.split()).strip().rstrip(".,")

def img_credit(base):
    """-> (license, info, reference) for one image base name."""
    if base in LYRYX_IMGS:
        return ("CC BY 4.0 (adapted from A First Course in Linear Algebra, Kuttler/Lyryx)",
                "Adapted from A First Course in Linear Algebra by Ken Kuttler (Lyryx Learning)",
                "Ken Kuttler, A First Course in Linear Algebra, Lyryx Learning, CC BY 4.0, "
                "https://lyryx.com/first-course-linear-algebra/")
    if base.startswith("dijkstra"):
        return (BOOK_LIC, "", "")
    if base == "feasiblePolytope":
        return ("CC BY 3.0 US",
                "Foundations of Applied Mathematics (BYU ACME), Volume 2, Simplex lab",
                "Foundations of Applied Mathematics, BYU ACME labs, CC BY 3.0 US, "
                "https://github.com/Foundations-of-Applied-Mathematics")
    e = bib_entry(base)
    if e:
        author = e.get("author", "")
        if author and not OWN.search(author):
            lics = re.findall(r"\[([^\]]+)\]", author)
            lic = "; ".join(lics) if lics else "see reference"
            author_clean = clean_author(re.sub(r"\s*\[[^\]]+\]", "", author))
            title = e.get("title", base)
            url = e.get("url", "").split("#")[0]
            info = f"{author_clean}: {title}"
            ref = f'{author_clean}, "{title}", {lic}' + (f", {url}" if url else "")
            return (lic, info, ref)
    return (BOOK_LIC, "", "")

# ---------------------------------------------------------------- epub scan
opf = open(os.path.join(EPUB_DIR, "OEBPS", "content.opf"), encoding="utf-8").read()
manifest = {}
for tag in re.findall(r"<item\b[^>]*>", opf):
    h = re.search(r"href=['\"]([^'\"]+)['\"]", tag)
    i = re.search(r"\bid=['\"]([^'\"]+)['\"]", tag)
    if h and i:
        manifest[i.group(1)] = h.group(1)
spine = [manifest[i] for i in re.findall(r"<itemref[^>]*idref=['\"]([^'\"]+)['\"]", opf)
         if i in manifest and manifest[i].endswith(".xhtml")]

CAP = re.compile(r'<(figcaption|div) class="caption"\s*>\s*<span class="id">'
                 r'(Figure|Table|Listing)[\s\xa0]*([A-Z0-9]+\.\d+)[^<]*</span>')
IMG = re.compile(r'<img[^>]*src="([^"]*)"[^>]*>')

def strip_tags(s):
    s = re.sub(r"<[^>]+>", "", s)
    s = html.unescape(s).replace("\xa0", " ").replace(" ", " ")
    return " ".join(s.split())

rows = []
for href in spine:
    path = os.path.join(EPUB_DIR, "OEBPS", href)
    if not os.path.exists(path):
        continue
    txt = open(path, encoding="utf-8").read()
    events = []
    for m in CAP.finditer(txt):
        events.append((m.start(), "cap", m))
    for m in IMG.finditer(txt):
        events.append((m.start(), "img", m))
    events.sort(key=lambda t: t[0])
    pending = []
    for _, kind, m in events:
        if kind == "img":
            src = m.group(1)
            base = os.path.splitext(os.path.basename(src))[0]
            if os.path.basename(src) in SKIP_IMGS:
                continue
            alt = re.search(r'alt="([^"]*)"', m.group(0))
            pending.append((base, strip_tags(alt.group(1)) if alt else ""))
        else:
            captype, num = m.group(2), m.group(3)
            if captype != "Figure":
                pending = []
                continue
            close = "</figcaption>" if m.group(1) == "figcaption" else "</div>"
            end = txt.find(close, m.end())
            caption = strip_tags(txt[m.end():end])
            alts = [a for _, a in pending if a]
            lic, infos, refs = [], [], []
            for base, _ in pending:
                l, i, r = img_credit(base)
                if l not in lic:
                    lic.append(l)
                if i and i not in infos:
                    infos.append(i)
                if r and r not in refs:
                    refs.append(r)
            special = [l for l in lic if l != BOOK_LIC]
            license_ = "; ".join(special) if special else BOOK_LIC
            rows.append({
                "num": num, "caption": caption,
                "alt": "\n".join(alts),
                "license": license_,
                "info": "; ".join(infos),
                "ref": "\n".join(refs),
                "imgs": [b for b, _ in pending],
            })
            pending = []

# ---------------------------------------------------------------- write xlsx
wb = openpyxl.load_workbook(TEMPLATE)
ws = wb["Sheet1"]
# wipe any leftover template rows
for r in range(ws.max_row, 1, -1):
    ws.delete_rows(r)
headers = ["Figure #", "Caption", "Alt Text", "License",
           "Info for figure reference", "Reference (correctly formatted)",
           "Notes", "AW notes"]
bold = Font(name="Arial", size=10, bold=True)
norm = Font(name="Arial", size=10)
wrap = Alignment(wrap_text=True, vertical="top")
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = bold; cell.alignment = wrap
for i, row in enumerate(rows, 2):
    vals = [row["num"], row["caption"], row["alt"], row["license"],
            row["info"], row["ref"], "", ""]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = norm; cell.alignment = wrap
        if c == 1:
            cell.number_format = "@"
widths = [10, 45, 60, 30, 35, 45, 20, 20]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
wb.save(OUT)
empty_alt = [r["num"] for r in rows if not r["alt"]]
print(f"rows: {len(rows)}; empty alt: {len(empty_alt)} {empty_alt}")
